import streamlit as st
import io
import zipfile
import pandas as pd
from pathlib import Path
from extractor_movimientos import parsear_archivo, crear_excel

# --- Page Config ---
st.set_page_config(
    page_title="Comprobantes a Excel",
    page_icon="📗",
    layout="centered"
)

# --- Styling ---
st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=Space+Mono:wght@400;700&family=Syne:wght@400;600;800&display=swap');

:root {
    --bg:        #0d0f14;
    --surface:   #141720;
    --border:    #252935;
    --accent:    #e8c84a;
    --accent2:   #4ae8a0;
    --text:      #e4e8f0;
    --muted:     #6b7280;
    --danger:    #f87171;
    --radius:    10px;
}

*, *::before, *::after { box-sizing: border-box; }

.stApp {
    background-color: var(--bg) !important;
    font-family: 'Syne', sans-serif;
    color: var(--text);
}

.block-container {
    padding-top: 2.5rem !important;
    padding-bottom: 3rem !important;
    max-width: 680px !important;
}

h1, h2, h3, h4, p, span, div, label {
    color: var(--text) !important;
}

/* Header */
.etl-logo {
    font-family: 'Space Mono', monospace;
    font-size: 0.7rem;
    letter-spacing: 0.35em;
    color: var(--accent) !important;
    text-transform: uppercase;
    text-align: center;
    margin-bottom: 0.8rem;
}
.etl-title {
    font-family: 'Syne', sans-serif !important;
    font-weight: 800;
    font-size: 2.8rem !important;
    line-height: 1.4;
    color: var(--text) !important;
    text-align: center;
    margin: 0 0 0.5rem !important;
}
.etl-title span { color: var(--accent) !important; }
.etl-subtitle {
    font-size: 0.85rem;
    color: var(--muted) !important;
    font-family: 'Space Mono', monospace;
    letter-spacing: 0.05em;
    text-align: center;
}
.divider {
    border: none;
    border-top: 1px solid var(--border);
    margin: 1.8rem 0;
}

/* Cards */
.card {
    background: var(--surface);
    border: 1px solid var(--border);
    border-radius: var(--radius);
    padding: 1.6rem 1.8rem;
    margin-bottom: 1.2rem;
    position: relative;
    overflow: hidden;
}
.card::before {
    content: '';
    position: absolute;
    top: 0; left: 0; right: 0;
    height: 2px;
    background: linear-gradient(90deg, var(--accent), transparent);
}
.card-label {
    font-family: 'Space Mono', monospace;
    font-size: 0.75rem;
    font-weight: 700;
    letter-spacing: 0.2em;
    color: var(--accent) !important;
    text-transform: uppercase;
    margin-bottom: 1rem;
}

/* File uploader */
[data-testid="stFileUploader"] > div,
[data-testid="stFileUploader"] > div > div,
[data-testid="stFileUploader"] section,
[data-testid="stFileUploader"] section > div,
[data-testid="stFileUploadDropzone"],
[data-testid="stFileDropzoneInstructions"],
.stFileUploader > div,
.stFileUploader section {
    background: #1a1d24 !important;
    background-color: #1a1d24 !important;
    border: 1.5px dashed var(--border) !important;
    border-radius: var(--radius) !important;
    transition: border-color 0.2s ease;
}
[data-testid="stFileUploader"] > div:hover,
[data-testid="stFileUploadDropzone"]:hover,
.stFileUploader > div:hover {
    border-color: var(--accent) !important;
}
.stFileUploader label, [data-testid="stFileUploader"] label {
    color: var(--muted) !important;
}
[data-testid="stFileUploader"] small,
[data-testid="stFileUploader"] span,
[data-testid="stFileDropzoneInstructions"] span,
[data-testid="stFileDropzoneInstructions"] small,
[data-testid="stFileDropzoneInstructions"] div {
    color: var(--muted) !important;
}
.stFileUploader button, [data-testid="stFileUploader"] button {
    background: var(--surface) !important;
    color: var(--accent) !important;
    border: 1px solid var(--border) !important;
    border-radius: 6px !important;
    font-family: 'Space Mono', monospace !important;
    font-size: 0.75rem !important;
}
.stFileUploader button:hover, [data-testid="stFileUploader"] button:hover {
    border-color: var(--accent) !important;
}

/* Checkbox */
.stCheckbox label span { color: var(--text) !important; }
[data-testid="stCheckbox"] > label > div {
    border-color: var(--border) !important;
}

/* Main action button */
.stButton > button {
    width: 100% !important;
    background: var(--accent) !important;
    color: #0a0c10 !important;
    border: none !important;
    border-radius: var(--radius) !important;
    font-family: 'Syne', sans-serif !important;
    font-weight: 800 !important;
    font-size: 1rem !important;
    letter-spacing: 0.08em;
    height: 3.2em !important;
    margin-top: 0.5rem;
    transition: all 0.18s ease !important;
    box-shadow: 0 0 20px rgba(232,200,74,0.15);
    text-shadow: none !important;
    -webkit-text-fill-color: #0a0c10 !important;
}
.stButton > button:hover {
    background: #f5d84e !important;
    box-shadow: 0 0 30px rgba(232,200,74,0.3) !important;
    transform: translateY(-1px);
}
.stButton > button:active { transform: translateY(0); }

/* Download button */
[data-testid="stDownloadButton"] > button {
    background: transparent !important;
    color: var(--accent2) !important;
    border: 1.5px solid var(--accent2) !important;
    border-radius: var(--radius) !important;
    font-family: 'Space Mono', monospace !important;
    font-size: 0.8rem !important;
    letter-spacing: 0.06em;
    width: 100% !important;
    height: 3em !important;
    margin-top: 0.8rem;
    transition: all 0.18s ease !important;
}
[data-testid="stDownloadButton"] > button:hover {
    background: rgba(74,232,160,0.08) !important;
    box-shadow: 0 0 20px rgba(74,232,160,0.2) !important;
}

/* Alerts */
[data-testid="stAlert"] {
    border-radius: var(--radius) !important;
}
.stSuccess > div {
    background: rgba(74,232,160,0.07) !important;
    border: 1px solid rgba(74,232,160,0.25) !important;
}
.stSuccess p, .stSuccess span, .stSuccess strong { color: var(--accent2) !important; }

.stError > div {
    background: rgba(248,113,113,0.07) !important;
    border: 1px solid rgba(248,113,113,0.3) !important;
}
.stError p, .stError span { color: var(--danger) !important; }

.stWarning > div {
    background: rgba(232,200,74,0.07) !important;
    border: 1px solid rgba(232,200,74,0.25) !important;
}
.stWarning p, .stWarning span { color: var(--accent) !important; }

.stInfo > div {
    background: rgba(99,122,255,0.07) !important;
    border: 1px solid rgba(99,122,255,0.3) !important;
}
.stInfo p, .stInfo span, .stInfo strong { color: #a5b4fc !important; }

/* Spinner */
.stSpinner > div { border-top-color: var(--accent) !important; }

/* Stats row */
.stats-row {
    display: flex;
    gap: 0.8rem;
    margin-top: 1rem;
}
.stat-chip {
    flex: 1;
    background: #0a0c10;
    border: 1px solid var(--border);
    border-radius: 8px;
    padding: 0.7rem 0.5rem;
    text-align: center;
}
.stat-chip .stat-val {
    font-family: 'Space Mono', monospace;
    font-size: 1.3rem;
    font-weight: 700;
    color: var(--accent) !important;
    display: block;
}
.stat-chip .stat-lbl {
    font-size: 0.65rem;
    letter-spacing: 0.1em;
    color: var(--muted) !important;
    text-transform: uppercase;
    display: block;
    margin-top: 0.2rem;
}

/* Scrollbar */
::-webkit-scrollbar { width: 5px; }
::-webkit-scrollbar-track { background: var(--bg); }
::-webkit-scrollbar-thumb { background: var(--border); border-radius: 99px; }

/* Navbar / Header bar */
header[data-testid="stHeader"],
.stAppHeader,
header.stAppHeader {
    background: #1a1d24 !important;
    background-color: #1a1d24 !important;
    border-bottom: 1px solid var(--border) !important;
}
header[data-testid="stHeader"] *,
.stAppHeader * {
    color: var(--accent) !important;
}

/* Selectbox */
[data-testid="stSelectbox"] > div > div,
.stSelectbox > div > div {
    background: #1a1d24 !important;
    background-color: #1a1d24 !important;
    border: 1.5px solid var(--border) !important;
    border-radius: var(--radius) !important;
    color: var(--accent) !important;
}
[data-testid="stSelectbox"] > div > div:hover,
.stSelectbox > div > div:hover {
    border-color: var(--accent) !important;
}
[data-testid="stSelectbox"] span,
[data-testid="stSelectbox"] div[data-baseweb="select"] span,
.stSelectbox span {
    color: var(--accent) !important;
}
[data-testid="stSelectbox"] svg {
    fill: var(--accent) !important;
}
[data-testid="stSelectbox"] label {
    color: var(--muted) !important;
}
/* Selectbox dropdown menu */
[data-baseweb="popover"],
[data-baseweb="menu"],
ul[role="listbox"] {
    background: #1a1d24 !important;
    background-color: #1a1d24 !important;
    border: 1px solid var(--border) !important;
}
ul[role="listbox"] li,
[data-baseweb="menu"] li {
    background: #1a1d24 !important;
    color: var(--text) !important;
}
ul[role="listbox"] li:hover,
[data-baseweb="menu"] li:hover,
ul[role="listbox"] li[aria-selected="true"],
[data-baseweb="menu"] li[aria-selected="true"] {
    background: var(--surface) !important;
    color: var(--accent) !important;
}

/* Radio buttons */
[data-testid="stRadio"] > div > div > label > div {
    background: #1a1d24 !important;
}

/* Footer */
.etl-footer {
    text-align: center;
    padding-top: 2rem;
    font-family: 'Space Mono', monospace;
    font-size: 0.62rem;
    color: var(--muted) !important;
    letter-spacing: 0.15em;
}
</style>
""", unsafe_allow_html=True)


# ─── Header ────────────────────────────────────────────────────────────────────
st.markdown("""
<div>
    <h1 class="etl-title">Movimientos a<span> Excel</span></h1>
    <p class="etl-subtitle">TXT  →  XLSX</p>
</div>
<hr class="divider">
""", unsafe_allow_html=True)


# ─── Selector de herramienta ────────────────────────────────────────────────────────────
TOOL_MOVIMIENTOS = "Extracción de Movimientos (.txt)"
TOOL_PORTAL_IVA = "Movimientos Portal IVA limpio (.zip)"

herramienta = st.selectbox(
    "Seleccioná la herramienta:",
    options=[TOOL_MOVIMIENTOS, TOOL_PORTAL_IVA],
    index=0,
)

if herramienta == TOOL_MOVIMIENTOS:
        # ─── Card 01: Archivo ──────────────────────────────────────────────────────────
        st.markdown('<div class="card"><div class="card-label">01 · Archivo fuente</div>', unsafe_allow_html=True)
        uploaded_file = st.file_uploader(
            "Arrastrá tu archivo o hacé click para seleccionarlo",
            type=["txt", "prn"],
            label_visibility="visible"
        )
        st.markdown('</div>', unsafe_allow_html=True)


        # ─── Card 02: Opciones ─────────────────────────────────────────────────────────
        st.markdown('<div class="card"><div class="card-label">02 · Opciones de exportación</div>', unsafe_allow_html=True)
        OPT_SOLO = "Solo Movimientos"
        OPT_AUXILIAR = "Exportar con columna Auxiliar"
        OPT_RESUMENES = "Incluir hojas de resumen"
        OPT_ARCA = "Cruce de comprobantes con ARCA (En desarrollo)"

        modo_export = st.radio(
            "Seleccioná el modo de exportación:",
            options=[OPT_SOLO, OPT_AUXILIAR, OPT_RESUMENES, OPT_ARCA],
            index=0,
            help="Solo se puede elegir una opción a la vez."
        )
        con_auxiliar = modo_export == OPT_AUXILIAR
        con_resumenes = modo_export == OPT_RESUMENES
        cruce_arca = modo_export == OPT_ARCA
        st.markdown('</div>', unsafe_allow_html=True)

        # ─── Card 02b: Archivo ARCA (condicional) ──────────────────────────────────────
        df_arca = None
        if cruce_arca:
            st.markdown('<div class="card"><div class="card-label">02b · Archivo ARCA (.zip)</div>', unsafe_allow_html=True)
            uploaded_arca = st.file_uploader(
                "Subí el .zip descargado de ARCA con los comprobantes",
                type=["zip"],
                label_visibility="visible",
                key="arca_zip"
            )
            if uploaded_arca:
                try:
                    with zipfile.ZipFile(io.BytesIO(uploaded_arca.getvalue())) as zf:
                        all_files = [f for f in zf.namelist() if not f.endswith('/')]
                        if all_files:
                            target_file = all_files[0]
                            with zf.open(target_file) as data_file:
                                raw = data_file.read()
                            csv_text = raw.decode('latin-1')
                            sep = ';' if csv_text.count(';') > csv_text.count(',') else ','
                            df_arca = pd.read_csv(
                                io.StringIO(csv_text), sep=sep, on_bad_lines='skip'
                            )
                            # Mapear códigos de comprobante ARCA a tipos del sistema (con letra)
                            ARCA_TIPO_MAP = {
                                # Facturas
                                1: 'FC A', 6: 'FC B', 11: 'FC C', 51: 'FC M',
                                19: 'FC', 22: 'FC', 195: 'FC T',
                                201: 'FC A', 206: 'FC B', 211: 'FC C',
                                # Notas de Débito
                                2: 'ND A', 7: 'ND B', 12: 'ND C', 52: 'ND M',
                                20: 'ND', 37: 'ND', 196: 'ND T',
                                45: 'ND A', 46: 'ND B', 47: 'ND C',
                                202: 'ND A', 207: 'ND B', 212: 'ND C',
                                # Notas de Crédito
                                3: 'NC A', 8: 'NC B', 13: 'NC C', 53: 'NC M',
                                21: 'NC', 38: 'NC', 90: 'NC', 197: 'NC T',
                                43: 'NC B', 44: 'NC C', 48: 'NC A',
                                203: 'NC A', 208: 'NC B', 213: 'NC C',
                                # Tique Factura
                                81: 'TF A', 82: 'TF B', 111: 'TF C', 118: 'TF M',
                                # Tique
                                83: 'TK', 109: 'TK C', 110: 'TK',
                                112: 'TK A', 113: 'TK B', 114: 'TK C',
                                115: 'TK A', 116: 'TK B', 117: 'TK C',
                                119: 'TK M', 120: 'TK M',
                            }
                            col_tipo = 'Tipo de Comprobante'
                            if col_tipo in df_arca.columns:
                                df_arca[col_tipo] = pd.to_numeric(df_arca[col_tipo], errors='coerce').astype('Int64')
                                df_arca[col_tipo] = df_arca[col_tipo].map(ARCA_TIPO_MAP).fillna(df_arca[col_tipo].astype(str))

                            # ── Limpieza de columnas ARCA ──────────────────────────────
                            # Renombrar columnas (usa partial match para encodings rotos)
                            RENAME_RULES = [
                                (['fecha', 'emisi'], 'Fecha'),
                                (['tipo', 'comprobante'], 'Comprobante'),
                                (['punto', 'venta'], 'PV'),
                                (['mero', 'comprobante'], 'Nro.'),
                                (['tipo', 'doc', 'vendedor'], 'Tipo Doc.'),
                                (['nro', 'doc', 'vendedor'], 'CUIT'),
                                (['denominaci', 'vendedor'], 'Razon Social'),
                                (['importe', 'total'], 'Total'),
                        (['moneda', 'original'], 'Moneda'),
                        (['tipo', 'cambio'], 'Tipo Cambio'),
                                (['importe', 'no', 'gravado'], 'No Gravado'),
                                (['importe', 'exento'], 'Exento'),
                                (['pagos', 'cta', 'otros'], 'Otras Perc.'),
                                (['percepciones', 'ingresos', 'brutos'], 'Perc IIBB'),
                                (['impuestos', 'municipales'], 'Impuestos Munic.'),
                                (['percepciones', 'pagos', 'cuenta', 'iva'], 'Perc. IVA'),
                                (['impuestos', 'internos'], 'Imp. Int.'),
                                (['importe', 'otros', 'tributos'], 'Otros. Trib.'),
                                (['neto', 'gravado', 'iva', '0'], 'IVA 0%'),
                                (['neto', 'gravado', 'iva', '21'], 'Gravado IVA 21'),
                                (['importe', 'iva', '21'], 'IVA 21'),
                                (['neto', 'gravado', 'iva', '27'], 'Gravado IVA 27'),
                                (['importe', 'iva', '27'], 'IVA 27'),
                                (['neto', 'gravado', 'iva', '10'], 'Gravado IVA 10,5'),
                                (['importe', 'iva', '10'], 'IVA 10,5'),
                                (['neto', 'gravado', 'iva', '2'], 'Gravado IVA 2,5'),
                                (['importe', 'iva', '2'], 'IVA 2,5'),
                                (['neto', 'gravado', 'iva', '5%'], 'Gravado IVA 5'),
                                (['importe', 'iva', '5%'], 'IVA 5'),
                            ]
                            rename_map = {}
                            for keywords, new_name in RENAME_RULES:
                                for c in df_arca.columns:
                                    cl = c.strip().lower()
                                    if all(k in cl for k in keywords) and c not in rename_map:
                                        rename_map[c] = new_name
                                        break
                            df_arca = df_arca.rename(columns=rename_map)

                            # Eliminar columnas no deseadas
                            DROP_KEYWORDS = [
                                ['dito', 'fiscal', 'computable'],
                                ['total', 'neto', 'gravado'],
                                ['total', 'iva'],
                                ['tipo', 'doc'],
                            ]
                            cols_to_drop = []
                            for keywords in DROP_KEYWORDS:
                                for c in df_arca.columns:
                                    cl = c.strip().lower()
                                    if all(k in cl for k in keywords):
                                        cols_to_drop.append(c)
                                        break
                            df_arca = df_arca.drop(columns=[c for c in cols_to_drop if c in df_arca.columns], errors='ignore')

                            # Mover Total al final
                            if 'Total' in df_arca.columns:
                                total_data = df_arca.pop('Total')
                                df_arca['Total'] = total_data

                            # Columna Auxiliar: Tipo + PV + Nro Comprobante + Nro Doc Vendedor
                            def find_col(df, keywords):
                                """Busca columna que contenga todas las keywords (case-insensitive)."""
                                for c in df.columns:
                                    cl = c.strip().lower()
                                    if all(k in cl for k in keywords):
                                        return c
                                return None

                            # Crear columna Auxiliar con nombres ya renombrados
                            aux_cols = ['Comprobante', 'PV', 'Nro.', 'CUIT']
                            if all(c in df_arca.columns for c in aux_cols):
                                df_arca['Auxiliar'] = (
                                    df_arca['Comprobante'].astype(str) +
                                    df_arca['PV'].astype(str) +
                                    df_arca['Nro.'].astype(str) +
                                    df_arca['CUIT'].astype(str)
                                )
                                # Mover Auxiliar justo antes de Total
                                cols = list(df_arca.columns)
                                cols.remove('Auxiliar')
                                total_pos = cols.index('Total') if 'Total' in cols else len(cols)
                                cols.insert(total_pos, 'Auxiliar')
                                df_arca = df_arca[cols]

                            # Columnas monetarias: desde 'No Gravado' en adelante (excluyendo Auxiliar)
                            all_cols = list(df_arca.columns)
                            ng_idx = all_cols.index('No Gravado') if 'No Gravado' in all_cols else None
                            if ng_idx is not None:
                                money_cols = [c for c in all_cols[ng_idx:] if c != 'Auxiliar']
                                for c in money_cols:
                                    # Convertir formato argentino: 1.234,56 -> 1234.56
                                    df_arca[c] = df_arca[c].astype(str).str.replace('.', '', regex=False).str.replace(',', '.', regex=False)
                                    df_arca[c] = pd.to_numeric(df_arca[c], errors='coerce').fillna(0)
                                # Eliminar columnas monetarias que son todo cero
                                empty_money = [c for c in money_cols if (df_arca[c] == 0).all()]
                                df_arca = df_arca.drop(columns=empty_money)

                            st.success(f"**{target_file}** · {len(df_arca)} comprobantes leídos de ARCA")
                        else:
                            st.error("El .zip está vacío")
                except Exception as e:
                    st.error(f"Error al leer el .zip: {str(e)}")
            else:
                st.info("Subí el archivo .zip de ARCA para continuar")
            st.markdown('</div>', unsafe_allow_html=True)


        if uploaded_file is not None:
            filename = Path(uploaded_file.name).stem
            st.success(f"**{uploaded_file.name}** listo para procesar")

            st.markdown('<div class="card"><div class="card-label">03 · Procesar</div>', unsafe_allow_html=True)

            if st.button("⬡  Procesar Archivo"):
                try:
                    with st.spinner("Analizando información..."):
                        content = uploaded_file.getvalue().decode("latin-1")
                        transacciones, meta = parsear_archivo(content=content)

                    if not transacciones:
                        st.error("No se encontraron transacciones. Verificá el formato del archivo.")
                    else:
                        with st.spinner("Generando Excel..."):
                            output = io.BytesIO()
                            crear_excel(transacciones, meta, output,
                                        con_resumenes=con_resumenes,
                                        con_auxiliar=con_auxiliar,
                                        cruce_arca=cruce_arca,
                                        df_arca=df_arca)
                            output.seek(0)

                        st.success("✓  Proceso completado con éxito")

                        # Stats chips
                        from collections import Counter
                        tipos = Counter(t['Tipo'] for t in transacciones)
                        st.markdown(f"""
                        <div class="stats-row">
                            <div class="stat-chip">
                                <span class="stat-val">{len(transacciones)}</span>
                                <span class="stat-lbl">Total</span>
                            </div>
                            <div class="stat-chip">
                                <span class="stat-val">{tipos.get('FC', 0)}</span>
                                <span class="stat-lbl">Facturas</span>
                            </div>
                            <div class="stat-chip">
                                <span class="stat-val">{tipos.get('NC', 0)}</span>
                                <span class="stat-lbl">Notas Cred.</span>
                            </div>
                            <div class="stat-chip">
                                <span class="stat-val">{tipos.get('ND', 0) + tipos.get('TF', 0) + tipos.get('TK', 0)}</span>
                                <span class="stat-lbl">Otros</span>
                            </div>
                        </div>
                        """, unsafe_allow_html=True)

                        st.info(
                            f"**{meta.get('tipo_reporte', 'N/A')}** · "
                            f"{meta.get('razon_social', 'Contribuyente')} · "
                            f"{meta.get('periodo', '')}"
                        )

                        st.download_button(
                            label="↓  Descargar Excel",
                            data=output,
                            file_name=f"{filename}_procesado.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            use_container_width=True,
                        )

                except Exception as e:
                    st.error(f"Error al procesar el archivo: {str(e)}")
                    st.exception(e)

            st.markdown('</div>', unsafe_allow_html=True)

        else:
            st.markdown("""
            <div style="
                text-align: center;
                padding: 2rem 1rem;
                font-family: 'Space Mono', monospace;
                font-size: 0.72rem;
                color: #6b7280;
                letter-spacing: 0.12em;
            ">
                ESPERANDO ARCHIVO · PASO 01
            </div>
            """, unsafe_allow_html=True)


else:
    # ───────────────────────────────────────────────────────────────────────────────
    # HERRAMIENTA: Portal IVA limpio
    # ───────────────────────────────────────────────────────────────────────────────
    st.markdown('<div class="card"><div class="card-label">01 · Archivo ARCA (.zip)</div>', unsafe_allow_html=True)
    uploaded_zip_iva = st.file_uploader(
        "Subí el .zip descargado del Portal IVA de ARCA",
        type=["zip"],
        label_visibility="visible",
        key="portal_iva_zip"
    )
    st.markdown('</div>', unsafe_allow_html=True)

    if uploaded_zip_iva:
        st.success(f"**{uploaded_zip_iva.name}** listo para procesar")
        st.markdown('<div class="card"><div class="card-label">02 · Procesar</div>', unsafe_allow_html=True)

        if st.button("⬡  Procesar ZIP"):
            try:
                with st.spinner("Leyendo archivo ARCA..."):
                    with zipfile.ZipFile(io.BytesIO(uploaded_zip_iva.getvalue())) as zf:
                        all_files = [f for f in zf.namelist() if not f.endswith('/')]
                        if not all_files:
                            st.error("El .zip está vacío")
                            st.stop()
                        target_file = all_files[0]
                        with zf.open(target_file) as data_file:
                            raw = data_file.read()

                    csv_text = raw.decode('latin-1')
                    sep = ';' if csv_text.count(';') > csv_text.count(',') else ','
                    df_iva = pd.read_csv(io.StringIO(csv_text), sep=sep, on_bad_lines='skip')

                    # Detectar tipo (Compras/Ventas), CUIT y periodo del nombre del zip
                    import re as _re
                    zip_name_raw = uploaded_zip_iva.name.upper()
                    es_ventas = 'VENTA' in zip_name_raw
                    es_compras = 'COMPRA' in zip_name_raw
                    tipo_portal = 'VENTAS' if es_ventas else ('COMPRAS' if es_compras else 'PORTAL IVA')

                    # Buscar CUIT (11 dígitos) y periodo (YYYYMM o YYYY-MM)
                    cuit_match = _re.search(r'(\d{11})', zip_name_raw)
                    cuit_portal = cuit_match.group(1) if cuit_match else ''
                    periodo_match = _re.search(r'(\d{4})(\d{2})(?!\d)', zip_name_raw)
                    if periodo_match:
                        meses = ['','Enero','Febrero','Marzo','Abril','Mayo','Junio',
                                 'Julio','Agosto','Septiembre','Octubre','Noviembre','Diciembre']
                        m_num = int(periodo_match.group(2))
                        periodo_portal = f"{meses[m_num]} {periodo_match.group(1)}" if 1 <= m_num <= 12 else ''
                    else:
                        periodo_portal = ''

                    # Mapear códigos de comprobante
                    ARCA_TIPO_MAP = {
                        1: 'FC A', 6: 'FC B', 11: 'FC C', 51: 'FC M',
                        19: 'FC', 22: 'FC', 195: 'FC T',
                        201: 'FC A', 206: 'FC B', 211: 'FC C',
                        2: 'ND A', 7: 'ND B', 12: 'ND C', 52: 'ND M',
                        20: 'ND', 37: 'ND', 196: 'ND T',
                        45: 'ND A', 46: 'ND B', 47: 'ND C',
                        202: 'ND A', 207: 'ND B', 212: 'ND C',
                        3: 'NC A', 8: 'NC B', 13: 'NC C', 53: 'NC M',
                        21: 'NC', 38: 'NC', 90: 'NC', 197: 'NC T',
                        43: 'NC B', 44: 'NC C', 48: 'NC A',
                        203: 'NC A', 208: 'NC B', 213: 'NC C',
                        81: 'TF A', 82: 'TF B', 111: 'TF C', 118: 'TF M',
                        83: 'TK', 109: 'TK C', 110: 'TK',
                        112: 'TK A', 113: 'TK B', 114: 'TK C',
                        115: 'TK A', 116: 'TK B', 117: 'TK C',
                        119: 'TK M', 120: 'TK M',
                    }

                    def find_col_iva(df, keywords):
                        for c in df.columns:
                            cl = c.strip().lower()
                            if all(k in cl for k in keywords):
                                return c
                        return None

                    col_tipo_iva = find_col_iva(df_iva, ['tipo', 'comprobante'])
                    if col_tipo_iva:
                        df_iva[col_tipo_iva] = pd.to_numeric(df_iva[col_tipo_iva], errors='coerce').astype('Int64')
                        df_iva[col_tipo_iva] = df_iva[col_tipo_iva].map(ARCA_TIPO_MAP).fillna(df_iva[col_tipo_iva].astype(str))

                    # Renombrar columnas (funciona para compras y ventas)
                    RENAME_RULES = [
                        (['fecha', 'emisi'], 'Fecha'),
                        (['tipo', 'comprobante'], 'Comprobante'),
                        (['punto', 'venta'], 'PV'),
                        (['mero', 'comprobante', 'hasta'], 'Nro. Hasta'),
                        (['mero', 'comprobante'], 'Nro.'),
                        (['tipo', 'doc'], 'Tipo Doc.'),
                        (['nro', 'doc', 'vendedor'], 'CUIT'),
                        (['nro', 'doc', 'comprador'], 'CUIT'),
                        (['denominaci', 'vendedor'], 'Razon Social'),
                        (['denominaci', 'comprador'], 'Razon Social'),
                        (['fecha', 'vencimiento'], 'Fecha Vto. Pago'),
                        (['importe', 'total'], 'Total'),
                        (['moneda', 'original'], 'Moneda'),
                        (['tipo', 'cambio'], 'Tipo Cambio'),
                        (['importe', 'no', 'gravado'], 'No Gravado'),
                        (['importe', 'exento'], 'Exento'),
                        (['pagos', 'cta', 'otros'], 'Otras Perc.'),
                        (['percepciones', 'ingresos', 'brutos'], 'Perc IIBB'),
                        (['impuestos', 'municipales'], 'Impuestos Munic.'),
                        (['percepciones', 'pagos', 'cuenta', 'iva'], 'Perc. IVA'),
                        (['percepci', 'no', 'categorizados'], 'Perc. No Cat.'),
                        (['impuestos', 'internos'], 'Imp. Int.'),
                        (['importe', 'otros', 'tributos'], 'Otros. Trib.'),
                        (['neto', 'gravado', 'iva', '0'], 'IVA 0%'),
                        (['neto', 'gravado', 'iva', '21'], 'Gravado IVA 21'),
                        (['importe', 'iva', '21'], 'IVA 21'),
                        (['neto', 'gravado', 'iva', '27'], 'Gravado IVA 27'),
                        (['importe', 'iva', '27'], 'IVA 27'),
                        (['neto', 'gravado', 'iva', '10'], 'Gravado IVA 10,5'),
                        (['importe', 'iva', '10'], 'IVA 10,5'),
                        (['neto', 'gravado', 'iva', '2'], 'Gravado IVA 2,5'),
                        (['importe', 'iva', '2'], 'IVA 2,5'),
                        (['neto', 'gravado', 'iva', '5%'], 'Gravado IVA 5'),
                        (['importe', 'iva', '5%'], 'IVA 5'),
                    ]
                    rename_map = {}
                    for keywords, new_name in RENAME_RULES:
                        for c in df_iva.columns:
                            cl = c.strip().lower()
                            if all(k in cl for k in keywords) and c not in rename_map:
                                rename_map[c] = new_name
                                break
                    df_iva = df_iva.rename(columns=rename_map)

                    # Eliminar columnas no deseadas
                    DROP_KW = [
                        ['dito', 'fiscal', 'computable'],
                        ['total', 'neto', 'gravado'],
                        ['total', 'iva'],
                        ['tipo', 'doc'],
                        ['nro.', 'hasta'],
                        ['fecha', 'vto'],
                    ]
                    cols_to_drop = []
                    for kws in DROP_KW:
                        for c in df_iva.columns:
                            cl = c.strip().lower()
                            if all(k in cl for k in kws):
                                cols_to_drop.append(c)
                                break
                    df_iva = df_iva.drop(columns=[c for c in cols_to_drop if c in df_iva.columns], errors='ignore')

                    # Mover Total al final
                    if 'Total' in df_iva.columns:
                        total_data = df_iva.pop('Total')
                        df_iva['Total'] = total_data

                    # Columnas monetarias: convertir y limpiar
                    all_cols_iva = list(df_iva.columns)
                    non_money = {'Fecha', 'Comprobante', 'PV', 'Nro.', 'CUIT', 'Razon Social', 'Moneda', 'Tipo Cambio'}
                    money_cols_iva = [c for c in all_cols_iva if c not in non_money and c in df_iva.select_dtypes(include='object').columns]
                    for c in money_cols_iva:
                        df_iva[c] = df_iva[c].astype(str).str.replace('.', '', regex=False).str.replace(',', '.', regex=False)
                        df_iva[c] = pd.to_numeric(df_iva[c], errors='coerce').fillna(0)
                    # Rellenar NaN restantes en columnas numéricas
                    for c in all_cols_iva:
                        if c not in non_money and df_iva[c].dtype in ('float64', 'int64'):
                            df_iva[c] = df_iva[c].fillna(0)
                    # Eliminar columnas monetarias todo cero
                    empty_cols = [c for c in all_cols_iva if c not in non_money and c in df_iva.columns and df_iva[c].dtype in ('float64', 'int64') and (df_iva[c] == 0).all()]
                    df_iva = df_iva.drop(columns=empty_cols)

                with st.spinner("Generando Excel..."):
                    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
                    from openpyxl.utils import get_column_letter

                    output = io.BytesIO()
                    with pd.ExcelWriter(output, engine='openpyxl') as writer:
                        df_iva.to_excel(writer, sheet_name=tipo_portal, index=False, startrow=5)
                        ws = writer.sheets[tipo_portal]
                        n_cols = len(df_iva.columns)

                        title_font = Font(bold=True, size=14, color='FFFFFF')
                        title_fill = PatternFill('solid', fgColor='2F5496')
                        header_font = Font(bold=True, size=10, color='FFFFFF')
                        header_fill = PatternFill('solid', fgColor='4472C4')
                        header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
                        center_align = Alignment(horizontal='center', vertical='center')
                        thin_border = Border(
                            left=Side(style='thin'), right=Side(style='thin'),
                            top=Side(style='thin'), bottom=Side(style='thin')
                        )
                        zebra_fill = PatternFill('solid', fgColor='D6E4F0')
                        money_fmt = '$#,##0.00'

                        ws.merge_cells(f'A1:{get_column_letter(n_cols)}1')
                        ws['A1'] = f'IVA {tipo_portal} - ARCA'
                        ws['A1'].font = title_font; ws['A1'].fill = title_fill
                        ws['A1'].alignment = center_align

                        ws.merge_cells(f'A2:{get_column_letter(n_cols)}2')
                        sub_parts = [p for p in [f'CUIT: {cuit_portal}' if cuit_portal else '', periodo_portal, f'{len(df_iva)} comprobantes'] if p]
                        ws['A2'] = ' | '.join(sub_parts)
                        ws['A2'].font = Font(bold=True, size=11, color='2F5496')
                        ws['A2'].alignment = center_align

                        col_list = list(df_iva.columns)
                        non_money_set = {'Fecha', 'Comprobante', 'PV', 'Nro.', 'CUIT', 'Razon Social', 'Auxiliar'}
                        money_indices = [i + 1 for i, c in enumerate(col_list) if c not in non_money_set and df_iva[c].dtype in ('float64', 'int64')]

                        for col_idx in range(1, n_cols + 1):
                            cell = ws.cell(row=6, column=col_idx)
                            cell.font = header_font; cell.fill = header_fill
                            cell.alignment = header_align; cell.border = thin_border

                        for row_idx in range(7, len(df_iva) + 7):
                            for col_idx in range(1, n_cols + 1):
                                cell = ws.cell(row=row_idx, column=col_idx)
                                cell.alignment = center_align
                                if col_idx in money_indices:
                                    cell.number_format = money_fmt
                            if (row_idx - 7) % 2 == 0:
                                for col_idx in range(1, n_cols + 1):
                                    ws.cell(row=row_idx, column=col_idx).fill = zebra_fill

                        # Fila TOTAL con fórmulas SUM
                        total_row = len(df_iva) + 7
                        col_list = list(df_iva.columns)
                        non_money_set2 = {'Fecha', 'Comprobante', 'PV', 'Nro.', 'CUIT', 'Razon Social', 'Moneda', 'Tipo Cambio'}
                        for col_idx in range(1, n_cols + 1):
                            cell = ws.cell(row=total_row, column=col_idx)
                            col_name = col_list[col_idx - 1] if col_idx - 1 < len(col_list) else ''
                            if col_name not in non_money_set2 and col_idx in money_indices:
                                letter = get_column_letter(col_idx)
                                cell.value = f'=SUM({letter}7:{letter}{total_row - 1})'
                                cell.number_format = money_fmt
                            elif col_idx == 1:
                                cell.value = 'TOTAL'
                            cell.font = Font(bold=True, size=10, color='FFFFFF')
                            cell.fill = PatternFill('solid', fgColor='2F5496')
                            cell.alignment = center_align

                        for col_idx in range(1, n_cols + 1):
                            max_len = max(
                                len(str(ws.cell(row=r, column=col_idx).value or ''))
                                for r in range(6, min(len(df_iva) + 7, 50))
                            )
                            letter = get_column_letter(col_idx)
                            ws.column_dimensions[letter].width = max(max_len + 3, 8)

                    output.seek(0)

                st.success("✓  Proceso completado con éxito")

                from collections import Counter
                tipos_iva = Counter(df_iva['Comprobante']) if 'Comprobante' in df_iva.columns else {}
                fc_count = sum(v for k, v in tipos_iva.items() if str(k).startswith('FC'))
                nc_count = sum(v for k, v in tipos_iva.items() if str(k).startswith('NC'))
                otros_count = len(df_iva) - fc_count - nc_count

                st.markdown(f"""
                <div class="stats-row">
                    <div class="stat-chip">
                        <span class="stat-val">{len(df_iva)}</span>
                        <span class="stat-lbl">Total</span>
                    </div>
                    <div class="stat-chip">
                        <span class="stat-val">{fc_count}</span>
                        <span class="stat-lbl">Facturas</span>
                    </div>
                    <div class="stat-chip">
                        <span class="stat-val">{nc_count}</span>
                        <span class="stat-lbl">Notas Cred.</span>
                    </div>
                    <div class="stat-chip">
                        <span class="stat-val">{otros_count}</span>
                        <span class="stat-lbl">Otros</span>
                    </div>
                </div>
                """, unsafe_allow_html=True)

                zip_name = Path(uploaded_zip_iva.name).stem
                st.download_button(
                    label="↓  Descargar Excel",
                    data=output,
                    file_name=f"{zip_name}_portal_iva.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True,
                )

            except Exception as e:
                st.error(f"Error al procesar: {str(e)}")
                st.exception(e)

        st.markdown('</div>', unsafe_allow_html=True)
    else:
        st.markdown("""
        <div style="
            text-align: center;
            padding: 2rem 1rem;
            font-family: 'Space Mono', monospace;
            font-size: 0.72rem;
            color: #6b7280;
            letter-spacing: 0.12em;
        ">
            ESPERANDO ARCHIVO .ZIP · PASO 01
        </div>
        """, unsafe_allow_html=True)
